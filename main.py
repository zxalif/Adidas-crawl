from bs4 import BeautifulSoup
from json import load, loads
import openpyxl
import os
import requests
import functools
import logging
import json
import time


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
root = 'https://shop.adidas.jp'
men_page_url = 'https://shop.adidas.jp/men/'


def masked_logging(mask="***", max_str_length=30):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            masked_args = [truncate_string(arg, max_str_length) if isinstance(arg, str) else arg for arg in args]
            masked_kwargs = {key: (truncate_string(val, max_str_length) if isinstance(val, str) else val) for key, val in kwargs.items()}
            logging.info(f"Calling function '{func.__name__}' with arguments: {masked_args}, kwargs: {masked_kwargs}")
            try:
                return func(*args, **kwargs)
            except Exception as e:
                logging.error(f"Error occurred in function '{func.__name__}': {e}")
        return wrapper
    return decorator


def truncate_string(s, max_length):
    if len(s) > max_length:
        return s[:max_length] + "..."
    return s


class ExcelWriter:
    headers = {}
    def __init__(self, filename):
        self.filename = filename
        if os.path.exists(filename):
            self.workbook = openpyxl.load_workbook(filename)
        else:
            self.workbook = openpyxl.Workbook()
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)
        self.worksheets = {ws.title: ws for ws in self.workbook.worksheets}

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.save()

    def add_worksheet(self, title):
        if title not in self.worksheets:
            worksheet = self.workbook.create_sheet(title=title)
            self.worksheets[title] = worksheet
            ExcelWriter.headers[title] = False
        else:
            worksheet = self.worksheets[title]
        return worksheet

    def write_data(self, worksheet, data):

        if not ExcelWriter.headers[worksheet]:
            ExcelWriter.headers[worksheet] = True
            headers = list(data.keys()) if isinstance(data, dict) else list(data[0].keys())
            self.worksheets[worksheet].append(headers)
        if isinstance(data, dict):
            self.worksheets[worksheet].append([data[header] for header in data.keys()])
        else:
            for row in data:
                self.worksheets[worksheet].append([row[header] for header in row.keys()])

    def save(self):
        self.workbook.save(self.filename)


@masked_logging(mask="***")
def get_content(url):
    # can be ip blocker, best to wait a bit
    time.sleep(0.15)
    response = requests.get(url)
    return response.content.decode()


@masked_logging(mask="***")
def get_list_of_items(content, tag, classes, single=False):
    if not content: return ''
    soup = content
    if isinstance(content, str):
        soup = BeautifulSoup(content, 'html.parser')

    finder = soup.find if single else soup.findAll
    return finder(tag, classes)


@masked_logging(mask="***")
def get_attributes(soup_item, attributes):
    return {attr: soup_item.get(attr) for attr in attributes}


@masked_logging(mask="***")
def get_attribute(soup_item, attribute):
    if not soup_item: return None
    return soup_item.get(attribute)


@masked_logging(mask="***")
def has_next(content):
    li = get_list_of_items(content, 'li', {'class': 'test-next'}, True)
    anchor = get_list_of_items(li, 'a', {}, True)
    href = get_attribute(anchor, 'href')
    if not href: return None
    return root + href


@masked_logging(mask="***")
def get_breadcrump_items(content):
    breadcrumps = get_list_of_items(content, 'li', {'class': 'breadcrumbListItem'})[1:]
    breadcrumps = [bc.text for bc in breadcrumps if bc]
    return '/'.join(breadcrumps)


@masked_logging(mask="***")
def get_category(content):
    category = get_list_of_items(content, 'span', {'class': 'test-categoryName'}, True)
    return category.text if category else ''


@masked_logging(mask="***")
def get_product_name(content):
    name = get_list_of_items(content, 'h1', {'class': 'test-itemTitle'}, True)
    return name.text if name else ''


@masked_logging(mask="***")
def get_product_price(content):
    price = get_list_of_items(content, 'div', {'class': 'test-articlePrice'}, True)
    # "," would be an issue in csv
    return price.text.replace(',', '') if price else ''

@masked_logging(mask="***")
def get_product_sizes(content):
    sizes = get_list_of_items(content, 'button', {'class': 'sizeSelectorListItemButton'})
    sizes = [size.text for size in sizes if size]
    return '/'.join(sizes)


@masked_logging(mask="***")
def get_product_default_images(content, as_json=''):
    images = get_list_of_items(content, 'img', {'class': 'test-image'})
    images = [get_attribute(image, 'src') for image in images if image]
    fmt = 'https://shop.adidas.jp{}'
    loader_images = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('article', {})
        .get('image', {})
        .get('details', [])
    )
    loader_images = [li.get('imageUrl', {}).get('large', '') for li in loader_images if li]
    images = [fmt.format(img) for img in (images + loader_images) if img and 'itemCard_dummy.jpg' not in img]
    return ','.join(images)


@masked_logging(mask="***")
def get_product_sense(content):
    sense = get_list_of_items(content, 'span', {'class': 'test-marker'}, True)
    sense = get_attribute(sense, 'class')
    if not sense: return ''
    record = 0
    for sc in sense:
        if sc.startswith('mod-marker_'):
            record = sc.strip('mod-marker_').replace('_', '.')
            break
    return record

@masked_logging(mask="***")
def get_related_products(as_json):
    related_articles = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('relatedArticles', [])
    )
    related = []
    for each in related_articles:
        data = {
            'name': each['name'],
            'code': each['code'],
            'price': each['price']['current']['withTax'],
            'image': each['image'],
            'url': 'https://shop.adidas.jp/products/' + each['code'] + '/',
        }
        related.append(data)
    return related

@masked_logging(mask="***")
def get_product_description(as_json):
    article = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('article', {})
        .get('description', {})
    )
    return article.get('messages', {})


@masked_logging(mask="***")
def get_product_kws(as_json):
    categories = (
        as_json
        .get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('page', {})
        .get('categories', [])
    )
    return ','.join([category['label'] for category in categories])


@masked_logging(mask="***")
def additional_review_data(code, model=None):
    url = 'https://adidasjp.ugc.bazaarvoice.com/7896-ja_jp/{}/reviews.djs'.format(model)
    data = {'format': 'embeddedhtml', 'productattribute_itemKcod': code, 'page': 1}
    response = requests.get(url, data).content.decode().split('\n')
    review_bar = response[6].replace('var materials=', '').replace('},', '').replace('\n', '').replace('\\', '').strip('n"')
    review = get_list_of_items(review_bar, 'span', {'class': 'BVRRNumber', 'itemprop': 'ratingValue'}, True)
    review = review.text if review else ''
    percentage = get_list_of_items(review_bar, 'span', {'class': 'BVRRBuyAgainPercentage'}, True)
    percentage = percentage.text if percentage else ''

    container = get_list_of_items(review_bar, 'div', {'class': 'BVRRSecondaryRatingsContainer'}, True)
    if not container:
        return review, percentage, '', '', '', ''
    fit = get_list_of_items(container, 'div', {'class': 'BVRRRatingFit'}, True)
    fit = get_list_of_items(fit, 'img', {'class': 'BVImgOrSprite'}, True)
    fit = get_attribute(fit, 'alt') if fit else ''

    length = get_list_of_items(container, 'div', {'class': 'BVRRRatingLength'}, True)
    length = get_list_of_items(length, 'img', {'class': 'BVImgOrSprite'}, True)
    length = get_attribute(length, 'alt') if length else ''

    quality = get_list_of_items(container, 'div', {'class': 'BVRRRatingQuality'}, True)
    quality = get_list_of_items(quality, 'img', {'class': 'BVImgOrSprite'}, True)
    quality = get_attribute(quality, 'alt') if quality else ''

    comfort = get_list_of_items(container, 'div', {'class': 'BVRRRatingComfort'}, True)
    comfort = get_list_of_items(comfort, 'img', {'class': 'BVImgOrSprite'}, True)
    comfort = get_attribute(comfort, 'alt') if comfort else ''
    return review, percentage, fit, length, quality, comfort


@masked_logging(mask="***")
def get_product_reviews(as_json):
    reviews = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('model', {})
        .get('review', {})
    )
    code = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('productIdInQuery', '')
    )
    # to get more than 10 review increate the page and iterate over it,
    # i have decided not do that
    model = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('article', {})
        .get('modelCode', '')
    )
    avg_rev, percentage, fit, length, quality, comfort = additional_review_data(code, model)
    reviews['avg_rev'] = avg_rev
    reviews['percentage'] = percentage
    reviews['fit'] = fit
    reviews['length'] = length
    reviews['quality'] = quality
    reviews['comfort'] = comfort
    return reviews

@masked_logging(mask="***")
def build_tale_size(tale_size):
    if not tale_size: return ''
    headers = tale_size['header']['0'].values()
    headers = ['measure'] + [header.get('value') for header in headers if header.get('value')]
    data = [','.join(headers)]
    for body in tale_size['body']:
        _data = ','.join(each.get('value') for each in tale_size['body'][body].values())
        data.append(_data)
    return '/'.join(data)

@masked_logging(mask="***")
def get_product_tale_of_size(as_json=None):
    code = (
        as_json.get('props', {})
        .get('pageProps', {})
        .get('apis', {})
        .get('pdpInitialProps', {})
        .get('detailApi', {})
        .get('product', {})
        .get('article', {})
        .get('modelCode', '')
    )
    # the api key can be generate with requests_html or selenium or by manual
    # url = 'https://services.virtusize.jp/product/check?apiKey={}&externalId={}&version=1'.format(code, api_key)
    url = 'https://shop.adidas.jp/f/v1/pub/size_chart/{}'.format(code)
    response = requests.get(url)
    if not response.ok:
        return {}
    tale_size = response.json().get('size_chart', {}).get('0', {})
    return build_tale_size(tale_size)

@masked_logging(mask="***")
def build_related_products(related_products, related_for):
    for rp in related_products:
        rp['related_for'] = related_for
    return related_products


@masked_logging(mask="***")
def build_reviews(reviews, related_for):
    for review in reviews:
        review['author'] = review['author']['name']
        review['reviewRating'] = review['reviewRating']['ratingValue']
        review['related_for'] = related_for
    return reviews


@masked_logging(mask="***")
def get_products_detail(*products):
    data = []
    for product in products:
        # full_product_url = root + get_attribute(product, 'href')
        full_product_url = product
        try:
            product_content = get_content(full_product_url)
        except:
            product_content = ''
        if not product_content: continue

        loader_data = get_list_of_items(product_content, 'script', {'id': '__NEXT_DATA__'}, True)
        # as_json = eval(loader_data.text, {'true': True, 'false': False, 'null': None}, {}) if loader_data else '{}'
        as_json = loads(loader_data.text) if loader_data else {}

        product_description = get_product_description(as_json)
        reviews = get_product_reviews(as_json)
        product_detail = {
            'url': full_product_url,
            'breadcrumps': get_breadcrump_items(product_content),
            'category': get_category(product_content),
            'name': get_product_name(product_content),
            'price': get_product_price(product_content),
            'sizes': get_product_sizes(product_content),
            'images': get_product_default_images(product_content, as_json),
            'sense': get_product_sense(product_content),
            'description_title': product_description['title'],
            'description_breads': '\n'.join(product_description['breads']),
            'description_main': product_description['mainText'],
            'kws': get_product_kws(as_json),
            'review_count': reviews['reviewCount'],
            'avg_rev': reviews['avg_rev'],
            'percentage': reviews['percentage'],
            'fit': reviews['fit'],
            'length': reviews['length'],
            'quality': reviews['quality'],
            'comfort': reviews['comfort'],
            'tale_size': get_product_tale_of_size(as_json),
        }
        with ExcelWriter(filename='addidas.xlsx') as ew:
            ew.write_data(
                'product_detail',
                product_detail)
            ew.write_data(
                'related_products',
                build_related_products(get_related_products(as_json), full_product_url)
            )
            ew.write_data(
                'reviews',
                build_reviews(reviews['reviewSeoLd'], full_product_url))

@masked_logging(mask="***")
def get_items_from_category(category):
    category = category.split('?')[-1] + '&limit=120'

    # can be increase with page=...
    url = 'https://shop.adidas.jp/f/v1/pub/product/list?' + category
    res = requests.get(url)
    if not res.ok:
        return {}
    res = res.json()
    product_url_list = ['https://shop.adidas.jp/products/{}/'.format(code) for code in res['articles_sort_list']]

    products = get_products_detail(*product_url_list)



@masked_logging(mask="***")
def main():
    men_page_content = get_content(men_page_url)
    category_of_men = get_list_of_items(men_page_content, 'a', {'class': 'lpc-teaserCarousel_link'})

    with ExcelWriter(filename='addidas.xlsx') as ew:
        ew.add_worksheet('product_detail')
        ew.add_worksheet('related_products')
        ew.add_worksheet('reviews')
    for category in category_of_men:
        href = get_attribute(category, 'href')

        # not category url for this category
        if not href:
            continue

        items = get_items_from_category(href)


if __name__ == '__main__':
    main()
